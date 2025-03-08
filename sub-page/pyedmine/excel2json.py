import os
import json
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def excel_to_json(input_dir="/Users/dream/myProjects/ZhijieXiong/sub-page/pyedmine/excel_data", 
                  output_dir="/Users/dream/myProjects/ZhijieXiong/sub-page/pyedmine/data"):
    """
    将 Excel 文件转换为 JSON 格式
    参数：
    input_dir: Excel 文件存放目录
    output_dir: JSON 输出目录
    """
    os.makedirs(output_dir, exist_ok=True)

    # 遍历所有 Excel 文件
    for filename in os.listdir(input_dir):
        if not filename.endswith((".xlsx", ".xls")):
            continue

        task_name = os.path.splitext(filename)[0]
        workbook = load_workbook(os.path.join(input_dir, filename))
        result = {"task": task_name, "scenes": {}}

        # 处理每个 sheet
        for sheet in workbook:
            scene_data = {"datasets": {}}
            merged_cells = sheet.merged_cells

            # ======================
            # 解析表头（数据集和指标）
            # ======================
            dataset_headers = []
            
            # 第一行：数据集名称（可能有合并单元格）
            for cell in sheet[1]:
                if cell.value is None:
                    continue
                
                # 查找合并单元格范围
                merged_range = next(
                    (mcr for mcr in merged_cells if cell.coordinate in mcr),
                    None
                )
                
                if merged_range:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    dataset_name = cell.value
                    dataset_headers.append({
                        "name": dataset_name,
                        "start_col": min_col,
                        "end_col": max_col
                    })
                else:
                    dataset_headers.append({
                        "name": cell.value,
                        "start_col": cell.column,
                        "end_col": cell.column
                    })

            # 第二行：指标名称
            metric_headers = [cell.value for cell in sheet[2] if cell.value]

            # ======================
            # 解析数据行
            # ======================
            for row in sheet.iter_rows(min_row=3):
                model_name = row[0].value
                if not model_name:
                    continue

                # 遍历每个数据集区域
                for dataset in dataset_headers:
                    dataset_name = dataset["name"]
                    start_col = dataset["start_col"]
                    end_col = dataset["end_col"]
                    
                    # 初始化数据集结构
                    if dataset_name not in scene_data["datasets"]:
                        scene_data["datasets"][dataset_name] = {"metrics": {}}
                    
                    # 遍历该数据集下的所有指标
                    for col_idx in range(start_col, end_col + 1):
                        metric_name = metric_headers[col_idx - 2]  # 列索引从1开始
                        cell_value = row[col_idx - 1].value  # 列索引从1开始
                        
                        # 处理特殊值
                        if isinstance(cell_value, str):
                            normalized_value = cell_value.lower()
                            if normalized_value in ["oom", "todo", "-", ""]:
                                final_value = cell_value
                            else:
                                try:
                                    final_value = float(cell_value)
                                except ValueError:
                                    final_value = cell_value
                        elif cell_value is None:
                            final_value = ""
                        else:
                            final_value = float(cell_value)
                        
                        # 写入指标数据
                        if metric_name not in scene_data["datasets"][dataset_name]["metrics"]:
                            scene_data["datasets"][dataset_name]["metrics"][metric_name] = {}
                        
                        scene_data["datasets"][dataset_name]["metrics"][metric_name][model_name] = final_value

            # 将当前 sheet 数据存入结果
            result["scenes"][sheet.title] = scene_data

        # 保存为 JSON
        output_path = os.path.join(output_dir, f"{task_name}.json")
        with open(output_path, "w") as f:
            json.dump(result, f, indent=2)

        print(f"Converted {filename} => {output_path}")

if __name__ == "__main__":
    excel_to_json()